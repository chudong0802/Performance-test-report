import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

#将指定文件筛选出来并转换为csv文件写到指定路径
#相对路径
os.system('mkdir chart')
os.system('mkdir xlsxfile')
path = '.\\analysis_module\\analysis_module\\'
out_path = './xlsxfile/'
all_file = os.listdir(path)
dealed_path = []
file_name = []

sheetname='Sheet1'
#选择xlsx文件
def choose_xlsx():
    for name in all_file:
        if name.startswith('com.') and os.path.splitext(name)[1] == '.xlsx'\
                or name.startswith('surface') or name.startswith('system'):
            new_name = os.path.splitext(name)[0]
            data_xlsx = pd.read_excel(path+name)
            data_xlsx.to_excel(out_path+new_name+'.xlsx',encoding='utf-8')


#处理xlsx文件
def deal_xlsx():
    for needed_name in os.listdir(out_path):
        needed_file = os.path.join(out_path,needed_name)
        dealed_path.append(needed_file)
        # print(dealed_path)
        filename = os.path.splitext(needed_name)[0]
        file_name.append(filename)

#遍历读取文件并画图
def deal_file_draw():
    for i in range(len(dealed_path)):
        dealed_file = pd.read_excel(dealed_path[i])
        dealed_filename = './chart/'+file_name[i]
        x = dealed_file['Folder']
        all_y = [dealed_file['dalvik max'], dealed_file['dalvik min'], dealed_file['dalvik avg']]
        filedir = os.makedirs(name=dealed_filename)
        plt.figure(figsize=(20, 10))
        # 坐标轴重合时，设置倾斜度
        # plt.xticks(rotation=45)
        plt.plot(x, all_y[0], color='orange', marker='s', label='max')
        plt.plot(x, all_y[1], color='green', marker='o', label='min')
        plt.plot(x, all_y[2], color='blue', marker='*', label='avg')
        for i in range(3):
            for a, b in zip(x, all_y[i]):
                plt.text(a, b+0.1, b, ha='center', va='bottom', fontsize=10)
        plt.locator_params('y', nbins=15)
        plt.legend(bbox_to_anchor=(0.9, 1), framealpha=0.2)
        plt.title('dalvik')
        image_path = dealed_filename + '/dalvik.png'
        plt.savefig(image_path)

        # bluetooth_native
        plt.figure(figsize=(20, 10))
        all_y1 = [dealed_file['native max'], dealed_file['native min'], dealed_file['native avg']]
        plt.plot(x, all_y1[0], color='orange', marker='s', label='max')
        plt.plot(x, all_y1[1], color='green', marker='o', label='min')
        plt.plot(x, all_y1[2], color='blue', marker='*', label='avg')
        for i in range(3):
            for a,b in zip(x,all_y1[i]):
                plt.text(a,b+0.2,b,ha='center',va='bottom',fontsize=10)
        plt.locator_params('y', nbins=15)
        plt.legend(bbox_to_anchor=(0.9, 1), framealpha=0.2)
        plt.title('native')

        image_path1 = dealed_filename + '/native.png'
        plt.savefig(image_path1)

        # bluetooth_cpu
        plt.figure(figsize=(20, 10))
        all_y2 = [dealed_file['cpu max'], dealed_file['cpu min'], dealed_file['cpu avg']]
        plt.plot(x, all_y2[0], color='orange', marker='s', label='max')
        plt.plot(x, all_y2[1], color='green', marker='o', label='min')
        plt.plot(x, all_y2[2], color='blue', marker='*', label='avg')
        for i in range(3):
            for a, b in zip(x, all_y2[i]):
                plt.text(a, b, b, ha='center', va='bottom', fontsize=10)
        plt.locator_params('y', nbins=15)
        plt.legend(bbox_to_anchor=(0.9, 1), framealpha=0.2)
        image_path2 = dealed_filename + '/cpu.png'
        plt.title('cpu')
        plt.savefig(image_path2)

        # 图片插入excel文件
        img = Image(image_path)
        img1 = Image(image_path1)
        img2 = Image(image_path2)
        wb = load_workbook(dealed_path[i])
        ws = wb[sheetname]
        rowsNum = ws.max_row
        # print(rowsNum)
        ws.add_image(img,'A{}'.format(rowsNum+2))
        wb.save(dealed_path[i])
        ws.add_image(img1, 'A{}'.format(rowsNum+62))
        wb.save(dealed_path[i])
        ws.add_image(img2, 'A{}'.format(rowsNum+122))
        wb.save(dealed_path[i])

if __name__ == '__main__':
    choose_xlsx()
    deal_xlsx()
    deal_file_draw()


